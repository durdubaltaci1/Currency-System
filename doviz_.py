import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QTableWidget, QTableWidgetItem, QDateEdit, QPushButton, QLabel, QComboBox, QFileDialog
from PyQt5.QtCore import Qt, QDate
import yfinance as yf
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class CurrencyTable(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Döviz Sistemi")
        self.setGeometry(100, 100, 1200, 600)  # Genişliği artırdık

        main_layout = QHBoxLayout()
        
        control_layout = QVBoxLayout()
        
        date_layout = QHBoxLayout()
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate().addYears(-5)) 

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDate(QDate.currentDate()) 

        self.currency_combo = QComboBox()
        self.currency_combo.addItems(["USD", "EUR"])

        self.fetch_button = QPushButton("Verileri Getir")
        self.fetch_button.clicked.connect(self.update_table)
        
        self.save_chart_button = QPushButton("Grafiği Kaydet")
        self.save_chart_button.clicked.connect(self.save_chart)

        self.save_excel_button = QPushButton("Excel'e Kaydet")
        self.save_excel_button.clicked.connect(self.save_to_excel)

        date_layout.addWidget(QLabel("Başlangıç Tarihi:"))
        date_layout.addWidget(self.start_date_edit)
        date_layout.addWidget(QLabel("Bitiş Tarihi:"))
        date_layout.addWidget(self.end_date_edit)
        date_layout.addWidget(QLabel("Para Birimi:"))
        date_layout.addWidget(self.currency_combo)
        date_layout.addWidget(self.fetch_button)
        date_layout.addWidget(self.save_chart_button)
        date_layout.addWidget(self.save_excel_button)

        control_layout.addLayout(date_layout)

        # Tablo widget'ı oluşturma
        self.table_widget = QTableWidget()
        control_layout.addWidget(self.table_widget)

        main_layout.addLayout(control_layout)

        # Grafik alanı
        self.figure = plt.figure()
        self.canvas = FigureCanvas(self.figure)
        main_layout.addWidget(self.canvas)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

        self.set_stylesheet()

    def set_stylesheet(self):
        self.table_widget.setStyleSheet("""
            QTableWidget {
                background-color: #f0f0f0;
                gridline-color: #d0d0d0;
                font: 14px "Arial";
            }
            QTableWidget::item {
                padding: 10px;
                border: 1px solid #d0d0d0;
            }
            QHeaderView::section {
                background-color: #4CAF50;
                color: white;
                padding: 4px;
                font: bold 14px "Arial";
                border: 1px solid #d0d0d0;
            }
            QTableWidget::item:selected {
                background-color: #4CAF50;
                color: white;
            }
        """)
        self.fetch_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px;
                font: bold 14px "Arial";
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.save_chart_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 10px;
                font: bold 14px "Arial";
            }
            QPushButton:hover {
                background-color: #1E88E5;
            }
        """)
        self.save_excel_button.setStyleSheet("""
            QPushButton {
                background-color: #FF5722;
                color: white;
                border: none;
                padding: 10px;
                font: bold 14px "Arial";
            }
            QPushButton:hover {
                background-color: #E64A19;
            }
        """)
        self.start_date_edit.setStyleSheet("QDateEdit { padding: 5px; font: 14px 'Arial'; }")
        self.end_date_edit.setStyleSheet("QDateEdit { padding: 5px; font: 14px 'Arial'; }")
        self.currency_combo.setStyleSheet("QComboBox { padding: 5px; font: 14px 'Arial'; }")

    def update_table(self):
        selected_currency = self.currency_combo.currentText()
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")

        # yfinance kullanarak döviz kurlarını alma
        if selected_currency == "USD":
            ticker = "TRY=X"  # USD/TRY kurunu al
        elif selected_currency == "EUR":
            ticker = "EURTRY=X"  # EUR/TRY kurunu al

        data = yf.download(ticker, start=start_date, end=end_date)
        if data.empty:
            print("Veri alınamadı.")
            return

        self.data = data  # Veriyi save_to_excel metodunda kullanmak için sakla

        # Tabloyu güncelleme
        self.table_widget.setRowCount(len(data))
        self.table_widget.setColumnCount(3)
        self.table_widget.setHorizontalHeaderLabels(["Tarih", "Değer", "Döviz Türü"])

        for row, (index, value) in enumerate(data['Close'].items()):
            date_str = index.strftime("%Y-%m-%d")
            self.table_widget.setItem(row, 0, QTableWidgetItem(date_str))
            self.table_widget.setItem(row, 1, QTableWidgetItem(str(value)))
            self.table_widget.setItem(row, 2, QTableWidgetItem(selected_currency))

        self.table_widget.resizeColumnsToContents()
        self.table_widget.resizeRowsToContents()

        # Grafik güncelleme
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.plot(data.index, data['Close'], label=selected_currency)
        ax.set_title(f"{selected_currency} Kuru")
        ax.set_xlabel("Tarih")
        ax.set_ylabel("Değer")
        ax.legend()
        self.canvas.draw()

    def save_chart(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Grafiği Kaydet", "", "PDF Files (*.pdf);;PNG Files (*.png);;JPG Files (*.jpg);;All Files (*)", options=options)
        if file_name:
            self.figure.savefig(file_name)
            print(f"Grafik kaydedildi: {file_name}")

    def save_to_excel(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Excel'e Kaydet", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_name:
            if not file_name.endswith('.xlsx'):
                file_name += '.xlsx'
            df = self.data[['Close']].reset_index()
            df.columns = ['Tarih', 'Değer']
            df['Döviz Türü'] = self.currency_combo.currentText()
            df.to_excel(file_name, index=False)

            # Excel dosyasını yükle ve tarih sütununu biçimlendir
            wb = load_workbook(file_name)
            ws = wb.active
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            for cell in ws["A"]:
                cell.number_format = 'YYYY-MM-DD'
            wb.save(file_name)
            print(f"Veriler Excel dosyasına kaydedildi: {file_name}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = CurrencyTable()
    window.show()
    sys.exit(app.exec_())
